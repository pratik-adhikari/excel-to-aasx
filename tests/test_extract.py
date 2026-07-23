"""Unit tests for workbook extraction and its audit metadata."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from excel_to_aasx.extract import extract_workbook, write_workbook_outputs


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _make_workbook(tmp_path: Path, sheets: dict[str, list[list]]) -> Path:
    """Write a minimal .xlsx file and return its path."""
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)
    path = tmp_path / "test.xlsx"
    wb.save(path)
    return path


HEADER_ROW = [
    "Element Name (idShort)",
    "Data Style / Field Type",
    "Universal Semantic ID (Global Key)",
    "Example Value",
    "Actual Value",
    "Unit",
]


# ──────────────────────────────────────────────────────────────────────────────
# Tests
# ──────────────────────────────────────────────────────────────────────────────

def test_extract_returns_expected_keys(tmp_path: Path) -> None:
    path = _make_workbook(tmp_path, {"Sheet1": [HEADER_ROW]})
    result = extract_workbook(path)
    assert "workbook" in result
    assert "source" in result
    assert "sheets" in result
    assert "Sheet1" in result["sheets"]


def test_extract_captures_workbook_filename(tmp_path: Path) -> None:
    path = _make_workbook(tmp_path, {"Sheet1": [HEADER_ROW]})
    result = extract_workbook(path)
    assert result["workbook"] == "test.xlsx"


def test_extract_includes_parsed_rows(tmp_path: Path) -> None:
    data_row = ["ManufacturerName", "MultiLanguageProperty",
                "0173-1#02-AAO677#002", "", "ACME Corp [en]", ""]
    path = _make_workbook(tmp_path, {"Sheet1": [HEADER_ROW, data_row]})
    result = extract_workbook(path)
    parsed = result["sheets"]["Sheet1"]["parsedRows"]
    # At least one row should have idShort = ManufacturerName
    id_shorts = [r.get("idShort") for r in parsed]
    assert "ManufacturerName" in id_shorts


def test_extract_multi_sheet(tmp_path: Path) -> None:
    path = _make_workbook(tmp_path, {
        "Alpha": [HEADER_ROW],
        "Beta":  [HEADER_ROW],
    })
    result = extract_workbook(path)
    assert set(result["sheets"].keys()) == {"Alpha", "Beta"}


def test_extract_empty_sheet_has_no_parsed_rows(tmp_path: Path) -> None:
    path = _make_workbook(tmp_path, {"Sheet1": [HEADER_ROW]})
    result = extract_workbook(path)
    parsed = result["sheets"]["Sheet1"]["parsedRows"]
    assert parsed == []


def test_write_workbook_outputs_creates_files(tmp_path: Path) -> None:
    path = _make_workbook(tmp_path, {"Sheet1": [HEADER_ROW]})
    workbook = extract_workbook(path)
    out_dir = tmp_path / "output"
    write_workbook_outputs(workbook, out_dir)

    # workbook.json must exist
    assert (out_dir / "test" / "workbook.json").exists()
    # per-sheet JSON for Sheet1 must exist
    sheet_files = list((out_dir / "test").glob("*.json"))
    assert len(sheet_files) == 2  # workbook.json + sheet1.json


def test_extract_actual_value_is_string(tmp_path: Path) -> None:
    """Actual values extracted from cells must be strings, not raw Python types."""
    data_row = ["", "SomeInteger", "Property", "", "", 42, ""]
    path = _make_workbook(tmp_path, {"Sheet1": [HEADER_ROW, data_row]})
    result = extract_workbook(path)
    parsed = result["sheets"]["Sheet1"]["parsedRows"]
    for row in parsed:
        actual = row.get("actualValue")
        if actual is not None:
            assert isinstance(actual, str), f"actualValue must be str, got {type(actual)}"


def test_extract_preserves_uncached_formula_and_warns(tmp_path: Path, capsys: pytest.CaptureFixture[str]) -> None:
    path = tmp_path / "formula.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(HEADER_ROW)
    ws.append(["Calculated", "Property", "", "", "", ""])
    ws["E2"] = "=1+1"
    wb.save(path)

    result = extract_workbook(path)
    cell = next(item for item in result["sheets"]["Sheet1"]["completeExtraction"]["cells"] if item["ref"] == "E2")

    assert cell["formula"] == "1+1"
    assert cell["value"] == ""
    assert "formula not pre-computed: Sheet1!E2" in capsys.readouterr().err
