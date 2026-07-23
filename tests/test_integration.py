"""Integration test: full extract → transform → validate pipeline.

Uses a minimal synthetic workbook (no real Schunk files required).
Covers the complete extraction-to-validation pipeline.
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from excel_to_aasx.extract import extract_workbook, write_workbook_outputs
from excel_to_aasx.validate import validate_aas_core_python


SCHUNK_CONFIG = Path("configs/companies/schunk.json")
REFERENCE_DIR = Path("third_party/admin-shell-io")

# Matches the real column name the extractor looks for
HEADER_ROW = [
    "Element Name (idShort)",
    "Data Style / Field Type",
    "Universal Semantic ID (Global Key)",
    "Example Value",
    "Actual Value",
    "Unit",
]


def _requires_schunk_config(request: pytest.FixtureRequest) -> None:
    if not SCHUNK_CONFIG.exists():
        pytest.skip("schunk company config not available")
    if not REFERENCE_DIR.exists():
        pytest.skip("vendored admin-shell-io template tree not available")


# ──────────────────────────────────────────────────────────────────────────────
# Fixtures
# ──────────────────────────────────────────────────────────────────────────────

@pytest.fixture()
def minimal_xlsx(tmp_path: Path) -> Path:
    """Create a minimal Excel workbook with the column structure the pipeline expects."""
    wb = openpyxl.Workbook()

    # Nameplate sheet (minimal mandatory row)
    ws = wb.active
    ws.title = "Digital Nameplate"
    ws.append(HEADER_ROW)
    ws.append(["ManufacturerName", "MultiLanguageProperty",
               "0173-1#02-AAO677#002", "", "ACME Corp [en]", ""])

    # Handover Documentation sheet
    ws2 = wb.create_sheet("Handover Documentation")
    ws2.append(HEADER_ROW)
    ws2.append(["DocumentDomainId", "Property", "", "", "DOC-001", ""])
    ws2.append(["DocumentClassId", "Property", "", "", "001", ""])

    # Technical Data sheet
    ws3 = wb.create_sheet("Technical Data")
    ws3.append(HEADER_ROW)
    ws3.append(["ManufacturerName", "MultiLanguageProperty",
                "0173-1#02-AAO677#002", "", "ACME Corp [en]", ""])

    # Carbon Footprint sheet (empty beyond header)
    ws4 = wb.create_sheet("Carbon Footprint")
    ws4.append(HEADER_ROW)

    # Maintenance Instructions sheet (empty beyond header)
    ws5 = wb.create_sheet("Maintenance Instructions")
    ws5.append(HEADER_ROW)

    xlsx_path = tmp_path / "test_workbook.xlsx"
    wb.save(xlsx_path)
    return xlsx_path


# ──────────────────────────────────────────────────────────────────────────────
# Tests
# ──────────────────────────────────────────────────────────────────────────────

def test_extract_produces_json(minimal_xlsx: Path, tmp_path: Path) -> None:
    """Step 1: extract_workbook should produce a JSON with the expected structure."""
    workbook = extract_workbook(minimal_xlsx)

    assert "workbook" in workbook
    assert "sheets" in workbook
    assert "Digital Nameplate" in workbook["sheets"]

    out_dir = tmp_path / "step1"
    write_workbook_outputs(workbook, out_dir)

    # extract.py slugifies filenames (underscores → hyphens)
    workbook_dirs = [d for d in out_dir.iterdir() if d.is_dir()]
    assert len(workbook_dirs) == 1
    workbook_dir = workbook_dirs[0]
    workbook_json = workbook_dir / "workbook.json"
    assert workbook_json.exists()

    loaded = json.loads(workbook_json.read_text())
    assert loaded["workbook"] == "test_workbook.xlsx"


def test_extract_handles_all_sheets(minimal_xlsx: Path) -> None:
    """extract_workbook should include all 5 sheets in its output."""
    workbook = extract_workbook(minimal_xlsx)
    assert set(workbook["sheets"].keys()) == {
        "Digital Nameplate",
        "Handover Documentation",
        "Technical Data",
        "Carbon Footprint",
        "Maintenance Instructions",
    }


def test_validate_aas_core_accepts_empty_environment() -> None:
    """An empty environment is accepted by the optional AAS Core baseline."""
    pytest.importorskip("aas_core3")
    result = validate_aas_core_python({})
    assert result["issueCounts"]["error"] == 0


@pytest.mark.skipif(
    not SCHUNK_CONFIG.exists() or not REFERENCE_DIR.exists(),
    reason="Schunk config / vendor templates not present",
)
def test_full_pipeline_extract_to_validation(
    minimal_xlsx: Path, tmp_path: Path
) -> None:
    """Full extract → transform → validate using the Schunk config and real templates."""
    from excel_to_aasx.company_config import load_company_config
    from excel_to_aasx.transform import build_workbook
    from excel_to_aasx.validate import validate_workbook

    config = load_company_config(SCHUNK_CONFIG)
    # This intentionally incomplete fixture opts into dummy values so the
    # integration test exercises the full pipeline rather than missing-data
    # failure behavior.
    config.setdefault("generationPolicy", {})["mandatoryMissingValue"] = "dummy"

    # Step 1
    workbook = extract_workbook(minimal_xlsx)
    step1_dir = tmp_path / "step1"
    write_workbook_outputs(workbook, step1_dir)

    # Step 2
    step2_dir = tmp_path / "step2"
    workbook_dirs = [d for d in step1_dir.iterdir() if d.is_dir()]
    assert len(workbook_dirs) == 1
    workbook_dir = workbook_dirs[0]
    build_workbook(workbook_dir, REFERENCE_DIR, step2_dir, config)

    # Step 3
    step3_dir = tmp_path / "step3"
    result = validate_workbook(
        step2_dir / workbook_dir.name,
        REFERENCE_DIR,
        step3_dir,
        aas_core_schema=None,
        aas_core_python=True,
        config=config,
    )

    # Must not crash; error count may be > 0 for a minimal workbook,
    # but the pipeline should complete and write a report.
    report_path = step3_dir / workbook_dir.name / "validation-report.json"
    assert report_path.exists(), "validation-report.json must be written"
    assert "issueCounts" in result
